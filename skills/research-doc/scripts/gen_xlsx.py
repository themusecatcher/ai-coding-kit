#!/usr/bin/env python3
"""
Excel 调研报告生成器
用法：python3 gen_xlsx.py --input content.json --output report.xlsx
"""

import argparse
import json
import sys
import os
from datetime import datetime

def check_deps():
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少 openpyxl，请执行: pip3 install --user openpyxl")
        sys.exit(1)

def create_xlsx(data: dict, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 样式定义 ──
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DATA_FONT = Font(name="Arial", size=10, color="333333")
    DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
    ALT_ROW_FILL = PatternFill(start_color="F7F8FA", end_color="F7F8FA", fill_type="solid")

    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1A73E8")
    SUBTITLE_FONT = Font(name="Arial", size=11, color="666666")

    THIN_BORDER = Border(
        left=Side(style="thin", color="D3D6DB"),
        right=Side(style="thin", color="D3D6DB"),
        top=Side(style="thin", color="D3D6DB"),
        bottom=Side(style="thin", color="D3D6DB"),
    )

    sections = data.get("sections", [])
    first_sheet = True

    # ── 概览页 ──
    ws = wb.active
    ws.title = "概览"

    ws["A1"] = data.get("title", "调研报告")
    ws["A1"].font = TITLE_FONT

    if data.get("subtitle"):
        ws["A2"] = data["subtitle"]
        ws["A2"].font = SUBTITLE_FONT

    row = 4
    meta_items = [
        ("日期", data.get("date", datetime.now().strftime("%Y-%m-%d"))),
    ]
    if data.get("author"):
        meta_items.insert(0, ("作者", data["author"]))

    for label, value in meta_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=value).font = Font(size=10)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="章节目录").font = Font(bold=True, size=11, color="1A73E8")
    row += 1
    for i, sec in enumerate(sections):
        ws.cell(row=row, column=1, value=f"{i + 1}. {sec.get('title', f'章节 {i + 1}')}")
        ws.cell(row=row, column=1).font = Font(size=10)
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # ── 各章节页 ──
    for i, section in enumerate(sections):
        sec_title = section.get("title", f"章节 {i + 1}")
        # Sheet 名称限制 31 字符
        sheet_name = f"{i + 1}_{sec_title}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        sec_type = section.get("type", "text")
        content = section.get("content", "")

        if sec_type == "table" and section.get("data"):
            # ── 表格数据 ──
            table_data = section["data"]
            headers = table_data.get("headers", [])
            rows = table_data.get("rows", [])

            # 标题行
            ws.cell(row=1, column=1, value=sec_title).font = TITLE_FONT
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))

            # 表头
            for j, h in enumerate(headers):
                cell = ws.cell(row=3, column=j + 1, value=str(h))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER

            # 数据行
            for r_idx, row in enumerate(rows):
                for c_idx, val in enumerate(row):
                    if c_idx < len(headers):
                        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=str(val))
                        cell.font = DATA_FONT
                        cell.alignment = DATA_ALIGNMENT
                        cell.border = THIN_BORDER
                        if r_idx % 2 == 1:
                            cell.fill = ALT_ROW_FILL

            # 自动列宽
            for j in range(len(headers)):
                col_letter = get_column_letter(j + 1)
                max_width = len(str(headers[j]))
                for row in rows:
                    if j < len(row):
                        max_width = max(max_width, len(str(row[j])))
                ws.column_dimensions[col_letter].width = min(max_width + 4, 50)

            # 冻结表头
            ws.freeze_panes = "A4"

        else:
            # ── 文本/列表数据 ──
            ws.cell(row=1, column=1, value=sec_title).font = TITLE_FONT

            lines = content.split("\n")
            row = 3
            for line in lines:
                stripped = line.strip()
                if stripped:
                    ws.cell(row=row, column=1, value=stripped).font = DATA_FONT
                    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                    row += 1

            ws.column_dimensions["A"].width = 80

    # 保存
    wb.save(output_path)
    sheet_count = len(wb.sheetnames)
    print(f"✅ Excel 已生成：{output_path}")
    print(f"   共 {sheet_count} 个 Sheet")


def main():
    parser = argparse.ArgumentParser(description="Excel 调研报告生成器")
    parser.add_argument("--input", required=True, help="内容 JSON 文件路径")
    parser.add_argument("--output", required=True, help="输出 XLSX 文件路径")
    args = parser.parse_args()

    check_deps()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    output_dir = os.path.dirname(args.output)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    create_xlsx(data, args.output)


if __name__ == "__main__":
    main()
